"""
Gera o template Excel padronizado para importação da Análise Mensal no dashboard OKR.
Uso: python criar_template_analise.py
Saída: template_analise_mensal.xlsx (na mesma pasta)
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

CHARCOAL = "343748"
NAVY     = "2C3799"
LGRAY    = "E7EBEE"
WHITE    = "FFFFFF"
DGRAY    = "2D2D2D"
AMBER    = "FAEEDA"
GREEN_BG = "EAF3DE"
RED_BG   = "FCEBEB"

def hdr(ws, row, col, text, bg=NAVY, fg=WHITE, bold=True, size=11):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def lbl(ws, row, col, text, bg=LGRAY, fg=DGRAY, bold=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", bold=bold, color=fg, size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(vertical="center")
    return c

def val(ws, row, col, value=None, fmt=None, note=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", size=10, color="0000FF")  # azul = input
    c.alignment = Alignment(vertical="center")
    if fmt:
        c.number_format = fmt
    return c

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_border(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row+1):
        for c in range(min_col, max_col+1):
            ws.cell(r, c).border = thin_border()

# ─────────────────────────────────────────────────────────
wb = Workbook()

# ══════════════════════════════════════════════════════════
# ABA 1 — CONFIG
# ══════════════════════════════════════════════════════════
ws_cfg = wb.active
ws_cfg.title = "CONFIG"
ws_cfg.sheet_view.showGridLines = False

ws_cfg.merge_cells("A1:D1")
c = ws_cfg["A1"]
c.value = "ANÁLISE MENSAL — CONFIGURAÇÃO"
c.font = Font(name="Calibri", bold=True, color=WHITE, size=13)
c.fill = PatternFill("solid", fgColor=CHARCOAL)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_cfg.row_dimensions[1].height = 30

ws_cfg.merge_cells("A2:D2")
c = ws_cfg["A2"]
c.value = "Preencha esta aba antes de qualquer outra. O Mês/Ano define a chave de importação no dashboard."
c.font = Font(name="Calibri", italic=True, color="666666", size=9)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_cfg.row_dimensions[2].height = 18

campos = [
    ("Mês de referência",  "Mai",   "Abreviação: Jan Fev Mar Abr Mai Jun Jul Ago Set Out Nov Dez"),
    ("Ano",                2026,    "Ex: 2026"),
    ("Fonte — Projetos",   "Hydra / CPJ-3C", "Sistema de origem dos dados de projetos"),
    ("Fonte — Horas Time", "Hydra / Todas categorias", "Sistema de origem das horas do time"),
    ("Fonte — Consultoria","Hydra / Categoria CONSULTORIA + CPJ-3C", "Sistema de origem dos dados de consultoria"),
    ("HS Vendidas base (Consultoria)", 781, "Total de horas contratadas no mês — denominador da taxa"),
    ("Meta consultoria (%)",           52.5, "Meta de utilização para o mês (%)"),
    ("Meta projetos (dias)",           39.6, "Meta de prazo médio de implementação para o mês"),
]

for i, (campo, default, instrucao) in enumerate(campos, start=4):
    ws_cfg.row_dimensions[i].height = 20
    lbl(ws_cfg, i, 1, campo)
    val(ws_cfg, i, 2, default)
    c = ws_cfg.cell(row=i, column=3, value=instrucao)
    c.font = Font(name="Calibri", italic=True, color="888888", size=9)
    c.alignment = Alignment(vertical="center")

# validação mês
dv_mes = DataValidation(type="list",
    formula1='"Jan,Fev,Mar,Abr,Mai,Jun,Jul,Ago,Set,Out,Nov,Dez"',
    allow_blank=False, showErrorMessage=True,
    errorTitle="Mês inválido", error="Use a abreviação: Jan Fev Mar ...")
ws_cfg.add_data_validation(dv_mes)
dv_mes.add(ws_cfg["B4"])

for col, w in zip([1,2,3], [28, 18, 50]):
    ws_cfg.column_dimensions[get_column_letter(col)].width = w

apply_border(ws_cfg, 4, 11, 1, 2)


# ══════════════════════════════════════════════════════════
# ABA 2 — PROJETOS
# ══════════════════════════════════════════════════════════
ws_p = wb.create_sheet("PROJETOS")
ws_p.sheet_view.showGridLines = False

ws_p.merge_cells("A1:F1")
c = ws_p["A1"]
c.value = "PROJETOS ENCERRADOS NO MÊS"
c.font = Font(name="Calibri", bold=True, color=WHITE, size=12)
c.fill = PatternFill("solid", fgColor=CHARCOAL)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_p.row_dimensions[1].height = 28

ws_p.merge_cells("A2:F2")
c = ws_p["A2"]
c.value = "Liste todos os projetos CPJ-3C encerrados (go live) no mês. Tipo: JORNADA ou UPGRADE. Status: ok / warn / bad."
c.font = Font(name="Calibri", italic=True, color="666666", size=9)
c.alignment = Alignment(horizontal="left", vertical="center")
ws_p.row_dimensions[2].height = 16

headers_p = ["Cliente", "Responsável", "Data Encerramento\n(dd/mm/aaaa)", "Tipo", "Dias", "Status"]
for col, h in enumerate(headers_p, 1):
    hdr(ws_p, 3, col, h, bg=NAVY)
ws_p.row_dimensions[3].height = 30

# 20 linhas de input
dv_tipo = DataValidation(type="list", formula1='"JORNADA,UPGRADE"', allow_blank=True)
dv_st_p = DataValidation(type="list", formula1='"ok,warn,bad"', allow_blank=True)
ws_p.add_data_validation(dv_tipo)
ws_p.add_data_validation(dv_st_p)

for r in range(4, 24):
    ws_p.row_dimensions[r].height = 18
    for col in range(1, 7):
        c = ws_p.cell(row=r, column=col)
        c.font = Font(name="Calibri", size=10, color="0000FF")
        c.alignment = Alignment(vertical="center")
        if col == 3:
            c.number_format = "DD/MM/AAAA"
        if col == 5:
            c.number_format = "0"
    dv_tipo.add(ws_p.cell(r, 4))
    dv_st_p.add(ws_p.cell(r, 6))
    apply_border(ws_p, r, r, 1, 6)

apply_border(ws_p, 3, 23, 1, 6)

# Bloco resumo — calculado por fórmula
row_res = 25
ws_p.merge_cells(f"A{row_res}:B{row_res}")
c = ws_p[f"A{row_res}"]
c.value = "RESUMO (calculado automaticamente)"
c.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
c.fill = PatternFill("solid", fgColor=CHARCOAL)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_p.row_dimensions[row_res].height = 20

resumo = [
    ("Total projetos",    '=COUNTA(A4:A23)'),
    ("Projetos JORNADA",  '=COUNTIF(D4:D23,"JORNADA")'),
    ("Projetos UPGRADE",  '=COUNTIF(D4:D23,"UPGRADE")'),
    ("Média geral (dias)",'=IFERROR(AVERAGEIF(E4:E23,">"&0,E4:E23),0)'),
    ("Média JORNADA",     '=IFERROR(AVERAGEIFS(E4:E23,D4:D23,"JORNADA"),0)'),
    ("Média UPGRADE",     '=IFERROR(AVERAGEIFS(E4:E23,D4:D23,"UPGRADE"),0)'),
]
for i, (nome, formula) in enumerate(resumo, start=row_res+1):
    ws_p.row_dimensions[i].height = 18
    lbl(ws_p, i, 1, nome)
    c = ws_p.cell(row=i, column=2, value=formula)
    c.font = Font(name="Calibri", size=10, color="000000")  # preto = fórmula
    c.number_format = "0.0"
    c.alignment = Alignment(horizontal="right", vertical="center")
    apply_border(ws_p, i, i, 1, 2)

widths_p = [40, 25, 18, 12, 8, 10]
for col, w in enumerate(widths_p, 1):
    ws_p.column_dimensions[get_column_letter(col)].width = w


# ══════════════════════════════════════════════════════════
# ABA 3 — HORAS_TIME
# ══════════════════════════════════════════════════════════
ws_h = wb.create_sheet("HORAS_TIME")
ws_h.sheet_view.showGridLines = False

ws_h.merge_cells("A1:D1")
c = ws_h["A1"]
c.value = "HORAS DO TIME NO MÊS"
c.font = Font(name="Calibri", bold=True, color=WHITE, size=12)
c.fill = PatternFill("solid", fgColor=CHARCOAL)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_h.row_dimensions[1].height = 28

lbl(ws_h, 2, 1, "Capacidade referência (h/mês por pessoa):")
val(ws_h, 2, 2, 128, "0")
c = ws_h.cell(row=2, column=3, value="80% do contratado · ajuste se necessário")
c.font = Font(name="Calibri", italic=True, color="888888", size=9)

headers_h = ["Colaborador", "Horas no mês", "% do total", "Observação (férias, afastamento...)"]
for col, h in enumerate(headers_h, 1):
    hdr(ws_h, 3, col, h, bg=NAVY)
ws_h.row_dimensions[3].height = 28

for r in range(4, 19):
    ws_h.row_dimensions[r].height = 18
    c_nome = ws_h.cell(row=r, column=1)
    c_nome.font = Font(name="Calibri", size=10, color="0000FF")
    c_nome.alignment = Alignment(vertical="center")

    c_h = ws_h.cell(row=r, column=2)
    c_h.font = Font(name="Calibri", size=10, color="0000FF")
    c_h.number_format = "0.00"
    c_h.alignment = Alignment(horizontal="right", vertical="center")

    c_pct = ws_h.cell(row=r, column=3, value=f"=IFERROR(B{r}/SUM($B$4:$B$18),\"\")")
    c_pct.font = Font(name="Calibri", size=10, color="000000")
    c_pct.number_format = "0.0%"
    c_pct.alignment = Alignment(horizontal="right", vertical="center")

    c_obs = ws_h.cell(row=r, column=4)
    c_obs.font = Font(name="Calibri", size=10, color="0000FF")
    c_obs.alignment = Alignment(vertical="center")

    apply_border(ws_h, r, r, 1, 4)

apply_border(ws_h, 3, 18, 1, 4)

# Total
r_tot = 19
ws_h.row_dimensions[r_tot].height = 20
lbl(ws_h, r_tot, 1, "TOTAL", bg=CHARCOAL, fg=WHITE)
c = ws_h.cell(row=r_tot, column=2, value="=SUM(B4:B18)")
c.font = Font(name="Calibri", bold=True, size=10, color="000000")
c.number_format = "0.00"
c.alignment = Alignment(horizontal="right", vertical="center")
c.fill = PatternFill("solid", fgColor=CHARCOAL)
c_100 = ws_h.cell(row=r_tot, column=3, value=1.0)
c_100.font = Font(name="Calibri", bold=True, size=10, color="000000")
c_100.number_format = "0.0%"
c_100.fill = PatternFill("solid", fgColor=CHARCOAL)
apply_border(ws_h, r_tot, r_tot, 1, 4)

widths_h = [35, 15, 12, 40]
for col, w in enumerate(widths_h, 1):
    ws_h.column_dimensions[get_column_letter(col)].width = w


# ══════════════════════════════════════════════════════════
# ABA 4 — CONSULTORIA
# ══════════════════════════════════════════════════════════
ws_c = wb.create_sheet("CONSULTORIA")
ws_c.sheet_view.showGridLines = False

ws_c.merge_cells("A1:E1")
c = ws_c["A1"]
c.value = "CONSULTORIA — UTILIZAÇÃO DE HORAS NO MÊS"
c.font = Font(name="Calibri", bold=True, color=WHITE, size=12)
c.fill = PatternFill("solid", fgColor=CHARCOAL)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_c.row_dimensions[1].height = 28

ws_c.merge_cells("A2:E2")
c = ws_c["A2"]
c.value = "Liste todos os clientes com horas executadas. Clientes com 0h no mês: preencha na seção abaixo."
c.font = Font(name="Calibri", italic=True, color="666666", size=9)
c.alignment = Alignment(horizontal="left", vertical="center")
ws_c.row_dimensions[2].height = 16

headers_c = ["Cliente", "HS Executadas", "Taxa Utilização (%)", "Consultor Responsável", "Status"]
for col, h in enumerate(headers_c, 1):
    hdr(ws_c, 3, col, h, bg=NAVY)
ws_c.row_dimensions[3].height = 28

dv_st_c = DataValidation(type="list", formula1='"ok,warn,bad,info"', allow_blank=True)
ws_c.add_data_validation(dv_st_c)

for r in range(4, 34):
    ws_c.row_dimensions[r].height = 18
    for col in [1, 4]:
        c = ws_c.cell(row=r, column=col)
        c.font = Font(name="Calibri", size=10, color="0000FF")
        c.alignment = Alignment(vertical="center")
    for col in [2, 3]:
        c = ws_c.cell(row=r, column=col)
        c.font = Font(name="Calibri", size=10, color="0000FF")
        c.number_format = "0.00"
        c.alignment = Alignment(horizontal="right", vertical="center")
    dv_st_c.add(ws_c.cell(r, 5))
    apply_border(ws_c, r, r, 1, 5)

apply_border(ws_c, 3, 33, 1, 5)

# Seção clientes 0h
r_zero = 35
ws_c.merge_cells(f"A{r_zero}:E{r_zero}")
c = ws_c[f"A{r_zero}"]
c.value = "CLIENTES COM 0 HORAS UTILIZADAS NO MÊS"
c.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
c.fill = PatternFill("solid", fgColor="A32D2D")
c.alignment = Alignment(horizontal="center", vertical="center")
ws_c.row_dimensions[r_zero].height = 22

headers_z = ["Cliente", "HS Contratadas (pacote)", "Consultor Responsável", "", ""]
for col, h in enumerate(headers_z, 1):
    hdr(ws_c, r_zero+1, col, h, bg="C0392B")
ws_c.row_dimensions[r_zero+1].height = 22

for r in range(r_zero+2, r_zero+12):
    ws_c.row_dimensions[r].height = 18
    for col in [1, 3]:
        c = ws_c.cell(row=r, column=col)
        c.font = Font(name="Calibri", size=10, color="0000FF")
        c.alignment = Alignment(vertical="center")
    c = ws_c.cell(row=r, column=2)
    c.font = Font(name="Calibri", size=10, color="0000FF")
    c.number_format = "0"
    c.alignment = Alignment(horizontal="right", vertical="center")
    apply_border(ws_c, r, r, 1, 3)

apply_border(ws_c, r_zero+1, r_zero+11, 1, 3)

# Resumo consultoria
r_res = r_zero + 13
ws_c.merge_cells(f"A{r_res}:B{r_res}")
c = ws_c[f"A{r_res}"]
c.value = "RESUMO (calculado automaticamente)"
c.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
c.fill = PatternFill("solid", fgColor=CHARCOAL)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_c.row_dimensions[r_res].height = 20

res_c = [
    ("Total HS Executadas",           f'=SUM(B4:B33)'),
    ("Clientes com execução",         f'=COUNTIF(B4:B33,">"&0)'),
    ("Clientes ≥80% utilização",      f'=COUNTIFS(C4:C33,">="&80,C4:C33,"<>"&"")'),
    ("% clientes ≥80%",               f'=IFERROR(COUNTIFS(C4:C33,">="&80)/COUNTIF(B4:B33,">"&0),0)'),
    ("Clientes 0h (seção abaixo)",    f'=COUNTA(A{r_zero+2}:A{r_zero+11})'),
]
for i, (nome, formula) in enumerate(res_c, start=r_res+1):
    ws_c.row_dimensions[i].height = 18
    lbl(ws_c, i, 1, nome)
    c = ws_c.cell(row=i, column=2, value=formula)
    c.font = Font(name="Calibri", size=10, color="000000")
    c.number_format = "0.00" if i - r_res - 1 in [0] else ("0.0%" if i - r_res - 1 == 3 else "0")
    c.alignment = Alignment(horizontal="right", vertical="center")
    apply_border(ws_c, i, i, 1, 2)

widths_c = [42, 16, 20, 28, 10]
for col, w in enumerate(widths_c, 1):
    ws_c.column_dimensions[get_column_letter(col)].width = w

# ──────────────────────────────────────────────────────────
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "template_analise_mensal.xlsx")
wb.save(out)
print(f"Template criado: {out}")
