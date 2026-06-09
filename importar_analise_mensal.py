"""
Converte o template_analise_mensal.xlsx preenchido em JSON para importação
no dashboard OKR.

Uso:
    python importar_analise_mensal.py <arquivo.xlsx>
    python importar_analise_mensal.py junho_2026.xlsx

Saída:
    analise_<mes>_<ano>.json  (na mesma pasta do script)
    Copie e importe pelo botão "📥 Análise Mensal" no dashboard.
"""
import sys
import json
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERRO: openpyxl não instalado. Execute: pip install openpyxl")
    sys.exit(1)

def ler_str(ws, row, col):
    v = ws.cell(row, col).value
    return str(v).strip() if v is not None else ""

def ler_num(ws, row, col, default=None):
    v = ws.cell(row, col).value
    if v is None or str(v).strip() == "":
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default

def processar(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    # ── CONFIG ──────────────────────────────────────────────
    cfg = wb["CONFIG"]
    mes      = ler_str(cfg, 4, 2)
    ano      = int(ler_num(cfg, 5, 2) or datetime.now().year)
    fonte_p  = ler_str(cfg, 6, 2)
    fonte_h  = ler_str(cfg, 7, 2)
    fonte_c  = ler_str(cfg, 8, 2)
    hs_vend  = ler_num(cfg, 9, 2, 0)
    meta_c   = ler_num(cfg, 10, 2, 0)
    meta_p   = ler_num(cfg, 11, 2, 0)
    chave    = f"{mes}/{ano}"

    # ── PROJETOS ────────────────────────────────────────────
    ws_p = wb["PROJETOS"]
    lista_p = []
    for r in range(4, 24):
        cliente = ler_str(ws_p, r, 1)
        if not cliente:
            continue
        resp  = ler_str(ws_p, r, 2)
        enc   = ws_p.cell(r, 3).value
        if hasattr(enc, "strftime"):
            enc = enc.strftime("%d/%m/%Y")
        elif enc:
            enc = str(enc).strip()
        else:
            enc = ""
        tipo  = ler_str(ws_p, r, 4).upper()
        dias  = ler_num(ws_p, r, 5)
        st    = ler_str(ws_p, r, 6).lower() or "warn"
        lista_p.append({"cliente": cliente, "resp": resp, "enc": enc,
                         "dias": int(dias) if dias else 0, "tipo": tipo, "st": st})

    total   = len(lista_p)
    jornada = sum(1 for p in lista_p if p["tipo"] == "JORNADA")
    upgrade = total - jornada
    dias_todos   = [p["dias"] for p in lista_p if p["dias"]]
    dias_jornada = [p["dias"] for p in lista_p if p["tipo"] == "JORNADA" and p["dias"]]
    dias_upgrade = [p["dias"] for p in lista_p if p["tipo"] == "UPGRADE" and p["dias"]]

    def media(lst): return round(sum(lst)/len(lst), 1) if lst else 0

    resumo_p = {
        "total": total, "jornada": jornada, "upgrade": upgrade,
        "mediaDias": media(dias_todos),
        "mediaJornada": media(dias_jornada),
        "mediaUpgrade": media(dias_upgrade),
        "metaMes": meta_p
    }

    # ── HORAS TIME ──────────────────────────────────────────
    ws_h = wb["HORAS_TIME"]
    capacidade = ler_num(ws_h, 2, 2, 128)
    lista_h = []
    for r in range(4, 19):
        colab = ler_str(ws_h, r, 1)
        if not colab:
            continue
        horas = ler_num(ws_h, r, 2, 0)
        obs   = ler_str(ws_h, r, 4)
        lista_h.append({"colab": colab, "horas": horas, "obs": obs})

    total_h = sum(x["horas"] for x in lista_h)
    for x in lista_h:
        x["pct"] = round(x["horas"] / total_h, 3) if total_h else 0

    # ── CONSULTORIA ─────────────────────────────────────────
    ws_c = wb["CONSULTORIA"]
    top_clientes = []
    for r in range(4, 34):
        cliente = ler_str(ws_c, r, 1)
        if not cliente:
            continue
        hs_exec  = ler_num(ws_c, r, 2)
        taxa     = ler_num(ws_c, r, 3)
        consultor= ler_str(ws_c, r, 4)
        st       = ler_str(ws_c, r, 5).lower() or "warn"
        if taxa is not None:
            taxa = taxa / 100 if taxa > 5 else taxa  # aceita % ou decimal
        top_clientes.append({"cliente": cliente, "hsExec": hs_exec or 0,
                              "taxa": taxa, "consultor": consultor, "st": st})

    # clientes 0h (seção separada)
    r_zero_start = 37  # linha 35 = header seção, 36 = header colunas, 37+ = dados
    clientes_0h = []
    for r in range(r_zero_start, r_zero_start + 10):
        cliente = ler_str(ws_c, r, 1)
        if not cliente:
            continue
        hs_cont  = int(ler_num(ws_c, r, 2) or 0)
        consultor= ler_str(ws_c, r, 3)
        clientes_0h.append({"cliente": cliente, "hsCont": hs_cont, "consultor": consultor})

    total_exec = sum(x["hsExec"] for x in top_clientes)
    taxa_geral = round(total_exec / hs_vend, 3) if hs_vend else 0
    cli_acima  = sum(1 for x in top_clientes if x["taxa"] is not None and x["taxa"] >= 0.8)
    pct_acima  = round(cli_acima / len(top_clientes), 3) if top_clientes else 0

    resumo_c = {
        "taxa": taxa_geral,
        "metaMes": round(meta_c / 100, 3) if meta_c > 1 else meta_c,
        "clientesAcima80": pct_acima,
        "clientes0h": len(clientes_0h)
    }

    # ── MONTAR JSON FINAL ────────────────────────────────────
    resultado = {
        chave: {
            "projetos": {
                "fonte": fonte_p,
                "resumo": resumo_p,
                "lista": lista_p
            },
            "horasTime": {
                "fonte": fonte_h,
                "capacidade": int(capacidade),
                "lista": lista_h
            },
            "consultoria": {
                "fonte": fonte_c,
                "resumo": resumo_c,
                "topClientes": top_clientes,
                "clientes0h": clientes_0h
            }
        }
    }

    nome_saida = f"analise_{mes.lower()}_{ano}.json"
    saida = os.path.join(os.path.dirname(os.path.abspath(__file__)), nome_saida)
    with open(saida, "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    print(f"\n✅ JSON gerado com sucesso!")
    print(f"   Chave: {chave}")
    print(f"   Projetos: {total} | Colaboradores: {len(lista_h)} | Clientes consultoria: {len(top_clientes)}")
    print(f"   Arquivo: {saida}")
    print(f"\n   Agora importe pelo botão '📥 Análise Mensal' no dashboard.")
    return saida

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python importar_analise_mensal.py <arquivo.xlsx>")
        sys.exit(1)
    processar(sys.argv[1])
